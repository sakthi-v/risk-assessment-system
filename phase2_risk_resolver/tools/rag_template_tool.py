"""
RAG Tool for Reading Excel Templates
This tool allows agents to read Excel template structures from uploaded files
"""

from crewai.tools import tool
import openpyxl
from typing import Dict, List, Any
import os


@tool("read_excel_template")
def read_excel_template(filepath: str) -> str:
    """
    Read structure from Excel template file
    
    Args:
        filepath: Name of Excel file (e.g., "Risk_Acceptance_Form_For_Agentic_AI_RM_Use_Case.xlsx")
    
    Returns:
        Template structure as text description that agents can understand
    
    Examples:
        read_excel_template("Risk_Acceptance_Form_For_Agentic_AI_RM_Use_Case.xlsx")
        read_excel_template("Risk_Acceptance_Tracker_For_Agentic_AI_RM_Use_Case.xlsx")
    """
    try:
        # Try different possible locations for the file
        possible_paths = [
            filepath,  # Try exact path first
            f'./templates/{filepath}',
            f'./templates/{filepath}',
            f'./{filepath}',
            f'Follow template_Use case v1.0.xlsx',  # Hardcoded follow-up template
            f'./Follow template_Use case v1.0.xlsx'
        ]
        
        wb = None
        used_path = None
        
        for path in possible_paths:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                used_path = path
                break
        
        if wb is None:
            return f"ERROR: Could not find file '{filepath}'. Please ensure it's uploaded to /mnt/user-data/uploads/"
        
        result = f"📋 Template File: {filepath}\n"
        result += f"📂 Location: {used_path}\n\n"
        
        # Read each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            result += f"{'='*80}\n"
            result += f"📄 Sheet: {sheet_name}\n"
            result += f"{'='*80}\n\n"
            
            # Read all rows with content (up to row 50)
            for row_idx in range(1, min(sheet.max_row + 1, 50)):
                row_data = []
                
                # Check columns A through T (1-20)
                for col_idx in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        # Clean the value
                        value = str(cell.value).strip()
                        row_data.append(f"{col_letter}: {value}")
                
                if row_data:
                    result += f"Row {row_idx:2d}: {' | '.join(row_data)}\n"
            
            result += "\n"
        
        wb.close()
        
        result += f"\n{'='*80}\n"
        result += "✅ Template structure read successfully!\n"
        result += f"{'='*80}\n"
        
        return result
    
    except Exception as e:
        return f"❌ Error reading template: {str(e)}\n\nPlease check:\n1. File exists in /mnt/user-data/uploads/\n2. File name is correct\n3. File is a valid Excel file"


@tool("search_rag_knowledge")
def search_rag_knowledge(query: str) -> str:
    """
    Search RAG system for risk management documentation and templates
    
    Args:
        query: Search query (e.g., "risk acceptance form requirements", "treatment plan guidelines")
    
    Returns:
        Relevant content from RAG knowledge base
    
    Examples:
        search_rag_knowledge("risk acceptance form structure")
        search_rag_knowledge("treatment plan implementation guidelines")
        search_rag_knowledge("risk approval workflow requirements")
    """
    try:
        # Try to import and use your RAG search function
        # Adjust this import based on your actual RAG implementation
        try:
            from phase2_risk_resolver.tools.rag_tool import search_knowledge_base
            results = search_knowledge_base(query)
            
            if results:
                response = f"📚 RAG Search Results for: '{query}'\n\n"
                response += f"Found {len(results)} relevant document(s):\n\n"
                response += "="*80 + "\n\n"
                
                for i, result in enumerate(results, 1):
                    response += f"Document {i}:\n{result}\n\n"
                    response += "="*80 + "\n\n"
                
                return response
            else:
                return f"No RAG results found for: '{query}'"
        
        except ImportError:
            # Fallback: Return built-in knowledge about risk management
            return f"""
📚 RAG Search: '{query}'

Since RAG system is not configured, using built-in knowledge:

RISK ACCEPTANCE FORM REQUIREMENTS:
- Form must include: Risk ID, Risk Description, Risk Category
- Acceptance justification must be detailed and business-focused
- Compensating controls must be specific and measurable
- Approval chain must include: Risk Owner, Risk Approver, and based on risk level (CISO, CIO, Board)
- Valid until date must be specified
- Monitoring plan must be defined

TREATMENT PLAN REQUIREMENTS:
- Must include specific actions for each selected control
- Timeline must be realistic based on priority (CRITICAL/HIGH/MEDIUM/LOW)
- Cost must be within budget constraints
- Success criteria must be measurable
- Implementation roadmap must have phases
- Expected risk reduction must be calculated

BEST PRACTICES:
- All forms should follow organizational templates
- Use templates from: Risk_Acceptance_Form_For_Agentic_AI_RM_Use_Case.xlsx
- Ensure all required fields are populated
- Validate data formats (emails, dates, etc.)
"""
    
    except Exception as e:
        return f"❌ RAG search error: {str(e)}\n\nUsing fallback knowledge base."


# Test function
if __name__ == "__main__":
    # Test reading template
    print("Testing RAG Tool...")
    result = read_excel_template("Risk_Acceptance_Form_For_Agentic_AI_RM_Use_Case.xlsx")
    print(result)